"""
audit_attachments.py
====================
Script de auditoría de adjuntos en Firebase Storage.
"""

import sys
import os
import datetime
from pathlib import Path
from typing import Optional

from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QComboBox,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QFileDialog,
    QMessageBox,
    QTextEdit,
    QGroupBox,
    QLineEdit,
    QProgressBar,
    QInputDialog,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor


class AttachmentAuditor(QMainWindow):
    """Ventana principal del auditor de adjuntos."""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("🔍 Auditor de Adjuntos - Firebase Storage")
        self.resize(1200, 800)
        
        # Variables
        self.db = None
        self.bucket = None
        self.companies = []
        self.current_company_id = None
        self.attachments_storage = []
        self.attachments_firestore = []
        
        self._build_ui()
    
    def _build_ui(self):
        """Construye la interfaz."""
        central = QWidget()
        self.setCentralWidget(central)
        
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(16)
        
        # === HEADER ===
        header_label = QLabel("🔍 Auditor de Adjuntos en Firebase Storage")
        header_label.setStyleSheet("font-size: 20px; font-weight: 700; color: #0F172A;")
        main_layout.addWidget(header_label)
        
        subtitle = QLabel("Compara adjuntos en Storage vs Firestore")
        subtitle.setStyleSheet("font-size: 13px; color: #64748B;")
        main_layout.addWidget(subtitle)
        
        # === GRUPO: CONFIGURACIÓN ===
        config_group = QGroupBox("⚙️ Configuración")
        config_group.setStyleSheet("""
            QGroupBox {
                font-weight: 700;
                font-size: 14px;
                color: #1E293B;
                border: 2px solid #E5E7EB;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        config_layout = QVBoxLayout(config_group)
        
        # Credenciales
        cred_row = QHBoxLayout()
        cred_label = QLabel("Credenciales Firebase:")
        cred_label.setMinimumWidth(150)
        cred_label.setStyleSheet("color: #0F172A;")
        cred_row.addWidget(cred_label)
        
        self.cred_path_input = QLineEdit()
        self.cred_path_input.setPlaceholderText("Selecciona el archivo JSON de credenciales...")
        self.cred_path_input.setReadOnly(True)
        cred_row.addWidget(self.cred_path_input, 1)
        
        btn_browse = QPushButton("📂 Seleccionar")
        btn_browse.clicked.connect(self._select_credentials)
        cred_row.addWidget(btn_browse)
        
        self.btn_connect = QPushButton("🔌 Conectar")
        self.btn_connect.clicked.connect(self._connect_firebase)
        self.btn_connect.setEnabled(False)
        cred_row.addWidget(self.btn_connect)
        
        config_layout.addLayout(cred_row)
        main_layout.addWidget(config_group)
        
        # === GRUPO: FILTROS ===
        filters_group = QGroupBox("🔍 Filtros")
        filters_group.setStyleSheet(config_group.styleSheet())
        filters_layout = QVBoxLayout(filters_group)
        
        # Fila 1: Empresa
        row1 = QHBoxLayout()
        
        lbl_company = QLabel("Empresa:")
        lbl_company.setMinimumWidth(100)
        lbl_company.setStyleSheet("color: #0F172A;")
        row1.addWidget(lbl_company)
        
        self.company_combo = QComboBox()
        self.company_combo.setMinimumWidth(300)
        self.company_combo.currentIndexChanged.connect(self._on_company_changed)
        row1.addWidget(self.company_combo, 1)
        
        row1.addStretch()
        filters_layout.addLayout(row1)
        
        # Fila 2: Año y Mes
        row2 = QHBoxLayout()
        
        lbl_year = QLabel("Año:")
        lbl_year.setMinimumWidth(100)
        lbl_year.setStyleSheet("color: #0F172A;")
        row2.addWidget(lbl_year)
        
        self.year_combo = QComboBox()
        current_year = datetime.date.today().year
        for y in range(current_year - 3, current_year + 2):
            self.year_combo.addItem(str(y))
        self.year_combo.setCurrentText(str(current_year))
        row2.addWidget(self.year_combo)
        
        lbl_month = QLabel("Mes:")
        lbl_month.setMinimumWidth(80)
        lbl_month.setStyleSheet("color: #0F172A;")
        row2.addWidget(lbl_month)
        
        self.month_combo = QComboBox()
        months = [
            "Todos", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]
        for month in months:
            self.month_combo.addItem(month)
        current_month = datetime.date.today().month
        self.month_combo.setCurrentIndex(current_month)
        row2.addWidget(self.month_combo)
        
        self.btn_scan = QPushButton("🔍 Escanear Adjuntos")
        self.btn_scan.clicked.connect(self._scan_attachments)
        self.btn_scan.setEnabled(False)
        row2.addWidget(self.btn_scan)
        
        row2.addStretch()
        filters_layout.addLayout(row2)
        
        main_layout.addWidget(filters_group)
        
        # === BARRA DE PROGRESO ===
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(True)
        main_layout.addWidget(self.progress_bar)
        
        # === GRUPO: ESTADÍSTICAS ===
        stats_group = QGroupBox("📊 Estadísticas")
        stats_group.setStyleSheet(config_group.styleSheet())
        stats_layout = QHBoxLayout(stats_group)
        
        # Tarjetas de estadísticas
        self.stat_storage = self._create_stat_card("En Storage", "0", "#3B82F6")
        self.stat_firestore = self._create_stat_card("En Firestore", "0", "#10B981")
        self.stat_match = self._create_stat_card("Coinciden", "0", "#15803D")
        self.stat_orphan = self._create_stat_card("Huérfanos", "0", "#DC2626")
        self.stat_missing = self._create_stat_card("Faltantes", "0", "#EA580C")
        
        stats_layout.addWidget(self.stat_storage["widget"])
        stats_layout.addWidget(self.stat_firestore["widget"])
        stats_layout.addWidget(self.stat_match["widget"])
        stats_layout.addWidget(self.stat_orphan["widget"])
        stats_layout.addWidget(self.stat_missing["widget"])
        
        main_layout.addWidget(stats_group)
        
        # === TABS: RESULTADOS ===
        results_label = QLabel("📋 Resultados del Escaneo:")
        results_label.setStyleSheet("font-weight: 700; font-size: 14px; color: #1E293B; margin-top: 8px;")
        main_layout.addWidget(results_label)
        
        # Tabla de resultados
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(6)
        self.results_table.setHorizontalHeaderLabels([
            "Archivo", "Factura", "Estado", "Tamaño", "Última Modificación", "Ruta"
        ])
        
        header = self.results_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        
        self.results_table.setAlternatingRowColors(True)
        self.results_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        main_layout.addWidget(self.results_table, 1)
        
        # === LOG ===
        log_label = QLabel("📝 Log:")
        log_label.setStyleSheet("font-weight: 600; color: #1E293B; margin-top: 8px;")
        main_layout.addWidget(log_label)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(150)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background-color: #F9FAFB;
                border: 1px solid #E5E7EB;
                border-radius: 6px;
                padding: 8px;
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 11px;
                color: #0F172A;
            }
        """)
        main_layout.addWidget(self.log_text)
        
        # === BOTONES ACCIÓN ===
        action_row = QHBoxLayout()
        action_row.addStretch()
        
        btn_export = QPushButton("📊 Exportar a Excel")
        btn_export.clicked.connect(self._export_excel)
        action_row.addWidget(btn_export)
        
        btn_clear = QPushButton("🗑️ Limpiar")
        btn_clear.clicked.connect(self._clear_results)
        action_row.addWidget(btn_clear)
        # En _build_ui(), en la sección de botones de acción:

        btn_investigate = QPushButton("🔎 Investigar Huérfanos")
        btn_investigate.clicked.connect(self._investigate_orphans)
        action_row.addWidget(btn_investigate)
        
        main_layout.addLayout(action_row)
        
        # ✅ ESTILOS GLOBALES CORREGIDOS
        self.setStyleSheet("""
            QMainWindow {
                background-color: #F8F9FA;
            }
            QPushButton {
                background-color: #3B82F6;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 600;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #2563EB;
            }
            QPushButton:disabled {
                background-color: #CBD5E1;
                color: #94A3B8;
            }
            
            QComboBox {
                background-color: white;
                border: 1px solid #D1D5DB;
                border-radius: 6px;
                padding: 6px 12px;
                font-size: 13px;
                color: #0F172A;
            }
            QComboBox:hover {
                border-color: #3B82F6;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 6px solid #64748B;
                margin-right: 8px;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                color: #0F172A;
                selection-background-color: #EFF6FF;
                selection-color: #1E40AF;
                border: 1px solid #D1D5DB;
                border-radius: 6px;
                padding: 4px;
            }
            QComboBox QAbstractItemView::item {
                padding: 6px 12px;
                color: #0F172A;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #EFF6FF;
                color: #1E40AF;
            }
            
            QLineEdit {
                background-color: white;
                border: 1px solid #D1D5DB;
                border-radius: 6px;
                padding: 6px 12px;
                font-size: 13px;
                color: #0F172A;
            }
            QLineEdit:focus {
                border-color: #3B82F6;
                border-width: 2px;
            }
            
            QTableWidget {
                background-color: white;
                alternate-background-color: #F9FAFB;
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                gridline-color: #E5E7EB;
                color: #0F172A;
            }
            QTableWidget::item {
                color: #0F172A;
                padding: 4px 8px;
            }
            QTableWidget::item:selected {
                background-color: #EFF6FF;
                color: #1E40AF;
            }
            QHeaderView::section {
                background-color: #F1F5F9;
                border: none;
                padding: 8px;
                font-weight: 700;
                font-size: 12px;
                color: #475569;
            }
            
            QLabel {
                color: #0F172A;
            }
        """)
    
    def _create_stat_card(self, title: str, value: str, color: str):
        """Crea una tarjeta de estadística."""
        widget = QWidget()
        widget.setStyleSheet(f"""
            QWidget {{
                background-color: white;
                border-left: 4px solid {color};
                border-radius: 8px;
                padding: 12px;
            }}
        """)
        
        layout = QVBoxLayout(widget)
        layout.setSpacing(4)
        
        title_label = QLabel(title)
        title_label.setStyleSheet("color: #64748B; font-size: 11px; font-weight: 600;")
        layout.addWidget(title_label)
        
        value_label = QLabel(value)
        value_label.setStyleSheet(f"color: {color}; font-size: 24px; font-weight: 700;")
        layout.addWidget(value_label)
        
        return {"widget": widget, "value": value_label}
    
    # ========================================
    # MÉTODOS DE CONEXIÓN
    # ========================================
    
    def _select_credentials(self):
        """Selecciona el archivo de credenciales."""
        fname, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Credenciales de Firebase",
            "",
            "JSON Files (*.json)"
        )
        
        if fname:
            self.cred_path_input.setText(fname)
            self.btn_connect.setEnabled(True)
    
    def _extract_bucket_name(self, cred_path: str) -> str:
        """Extrae el nombre del bucket desde las credenciales."""
        import json
        
        try:
            with open(cred_path, 'r') as f:
                data = json.load(f)
                project_id = data.get('project_id', '')
                
                # ✅ CORRECCIÓN: Usar .firebasestorage.app
                return f"{project_id}.firebasestorage.app"
        except Exception as e:
            print(f"⚠️ Error extrayendo bucket: {e}")
            return ""
    
    def _connect_firebase(self):
        """Conecta a Firebase."""
        cred_path = self.cred_path_input.text()
        
        if not cred_path or not os.path.exists(cred_path):
            QMessageBox.warning(self, "Error", "Selecciona un archivo de credenciales válido.")
            return
        
        self._log("🔌 Conectando a Firebase...")
        QApplication.processEvents()
        
        try:
            import firebase_admin
            from firebase_admin import credentials, firestore, storage
            
            # Obtener bucket name
            bucket_name = self._extract_bucket_name(cred_path)
            
            if not bucket_name:
                bucket_name, ok = QInputDialog.getText(
                    self,
                    "Nombre del Bucket",
                    "No se pudo detectar el bucket automáticamente.\n\n"
                    "Ingresa el nombre del bucket de Storage\n"
                    "(ejemplo: proyecto.firebasestorage.app):",
                    text="facot-app.firebasestorage.app"
                )
                
                if not ok or not bucket_name:
                    self._log("❌ Conexión cancelada")
                    return
            
            self._log(f"📦 Usando bucket: {bucket_name}")
            QApplication.processEvents()
            
            # Inicializar Firebase
            if not firebase_admin._apps:
                cred = credentials.Certificate(cred_path)
                firebase_admin.initialize_app(cred, {
                    'storageBucket': bucket_name
                })
            
            self.db = firestore.client()
            self.bucket = storage.bucket()
            
            self._log("✅ Firebase inicializado")
            QApplication.processEvents()
            
            # Verificar bucket
            try:
                test_blobs = list(self.bucket.list_blobs(max_results=1))
                self._log("✅ Bucket verificado y accesible")
            except Exception as e:
                self._log(f"⚠️ Advertencia: {e}")
                reply = QMessageBox.question(
                    self,
                    "Bucket no verificado",
                    f"No se pudo verificar el bucket.\n\n{e}\n\n¿Continuar?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply != QMessageBox.StandardButton.Yes:
                    return
            
            self._log("✅ Conectado exitosamente")
            
            # Cargar empresas
            self._load_companies()
            
            self.btn_connect.setEnabled(False)
            self.btn_scan.setEnabled(True)
            
        except Exception as e:
            self._log(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"No se pudo conectar:\n{e}")
    
    def _load_companies(self):
        """Carga las empresas desde Firestore."""
        if not self.db:
            return
        
        self._log("📊 Cargando empresas...")
        
        try:
            companies_ref = self.db.collection('companies')
            docs = companies_ref.stream()
            
            self.companies = []
            self.company_combo.clear()
            
            for doc in docs:
                data = doc.to_dict()
                company_name = data.get('name', doc.id)
                self.companies.append({
                    'id': doc.id,
                    'name': company_name,
                })
                self.company_combo.addItem(company_name, doc.id)
            
            self._log(f"✅ {len(self.companies)} empresas cargadas")
            
        except Exception as e:
            self._log(f"❌ Error cargando empresas: {e}")
    
    def _on_company_changed(self):
        """Evento al cambiar de empresa."""
        self.current_company_id = self.company_combo.currentData()
    
    # ========================================
    # ESCANEO DE ADJUNTOS
    # ========================================
    
    def _scan_attachments(self):
        """Escanea adjuntos en Storage y Firestore."""
        if not self.bucket or not self.db or not self.current_company_id:
            QMessageBox.warning(self, "Error", "Conecta a Firebase y selecciona una empresa.")
            return
        
        company_name = self.company_combo.currentText()
        year = self.year_combo.currentText()
        month_name = self.month_combo.currentText()
        
        # Convertir mes a número
        month = None
        if month_name != "Todos":
            months_map = {
                "Enero": "01", "Febrero": "02", "Marzo": "03", "Abril": "04",
                "Mayo": "05", "Junio": "06", "Julio": "07", "Agosto": "08",
                "Septiembre": "09", "Octubre": "10", "Noviembre": "11", "Diciembre": "12"
            }
            month = months_map.get(month_name)
        
        self._log("")
        self._log("="*80)
        self._log(f"🔍 ESCANEANDO ADJUNTOS")
        self._log(f"Empresa: {company_name}")
        self._log(f"Periodo: {month_name} {year}")
        self._log("="*80)
        
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.btn_scan.setEnabled(False)
        
        QApplication.processEvents()
        
        try:
            # ========================================
            # 1. ESCANEAR STORAGE
            # ========================================
            self._log("\n📁 Escaneando Firebase Storage...")
            self.progress_bar.setValue(20)
            QApplication.processEvents()
            
            # Normalizar nombre de empresa
            safe_company = (
                "".join(c for c in company_name if c.isalnum() or c in (" ", "-", "_"))
                .strip()
                .replace(" ", "_")
            ) or "company"
            
            prefix = f"Adjuntos/{safe_company}/{year}/"
            if month:
                prefix += f"{month}/"
            
            self._log(f"   Prefijo: {prefix}")
            
            blobs = self.bucket.list_blobs(prefix=prefix)
            
            self.attachments_storage = []
            for blob in blobs:
                self.attachments_storage.append({
                    'name': blob.name,
                    'size': blob.size,
                    'updated': blob.updated,
                    'content_type': blob.content_type,
                })
            
            self._log(f"   ✅ {len(self.attachments_storage)} archivos encontrados en Storage")
            
            # ========================================
            # 2. ESCANEAR FIRESTORE
            # ========================================
            self._log("\n📊 Escaneando Firestore...")
            self.progress_bar.setValue(60)
            QApplication.processEvents()
            
            invoices_ref = self.db.collection('invoices')
            query = invoices_ref.where('company_id', '==', self.current_company_id)
            
            docs = query.stream()
            
            self.attachments_firestore = []
            for doc in docs:
                data = doc.to_dict()
                
                # Filtrar por fecha
                invoice_date = data.get('invoice_date')
                if month and invoice_date:
                    try:
                        if hasattr(invoice_date, 'year'):
                            inv_year = invoice_date.year
                            inv_month = invoice_date.month
                        else:
                            date_str = str(invoice_date)[:10]
                            inv_year = int(date_str[:4])
                            inv_month = int(date_str[5:7])
                        
                        if inv_year != int(year) or inv_month != int(month):
                            continue
                    except:
                        pass
                
                storage_path = data.get('attachment_storage_path') or data.get('storage_path')
                
                if storage_path:
                    self.attachments_firestore.append({
                        'invoice_id': doc.id,
                        'invoice_number': data.get('invoice_number', 'N/A'),
                        'storage_path': storage_path,
                        'invoice_date': invoice_date,
                    })
            
            self._log(f"   ✅ {len(self.attachments_firestore)} facturas con adjunto en Firestore")
            
            # ========================================
            # 3. COMPARAR Y ANALIZAR
            # ========================================
            self._log("\n🔍 Comparando resultados...")
            self.progress_bar.setValue(80)
            QApplication.processEvents()
            
            self._analyze_results()
            
            self.progress_bar.setValue(100)
            self._log("\n✅ Escaneo completado")
            
        except Exception as e:
            self._log(f"\n❌ Error durante el escaneo: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            self.btn_scan.setEnabled(True)
            self.progress_bar.setVisible(False)
    
    def _analyze_results(self):
        """Analiza y compara los resultados."""
        storage_paths = set(att['name'] for att in self.attachments_storage)
        firestore_paths = set(att['storage_path'] for att in self.attachments_firestore)
        
        total_storage = len(storage_paths)
        total_firestore = len(firestore_paths)
        matches = len(storage_paths & firestore_paths)
        orphans = len(storage_paths - firestore_paths)
        missing = len(firestore_paths - storage_paths)
        
        self.stat_storage["value"].setText(str(total_storage))
        self.stat_firestore["value"].setText(str(total_firestore))
        self.stat_match["value"].setText(str(matches))
        self.stat_orphan["value"].setText(str(orphans))
        self.stat_missing["value"].setText(str(missing))
        
        self._log(f"\n📊 ESTADÍSTICAS:")
        self._log(f"   En Storage: {total_storage}")
        self._log(f"   En Firestore: {total_firestore}")
        self._log(f"   Coinciden: {matches}")
        self._log(f"   Huérfanos: {orphans}")
        self._log(f"   Faltantes: {missing}")
        
        self._populate_results_table()
    
    def _populate_results_table(self):
        """Pobla la tabla de resultados."""
        self.results_table.setRowCount(0)
        
        firestore_map = {att['storage_path']: att for att in self.attachments_firestore}
        storage_map = {att['name']: att for att in self.attachments_storage}
        
        all_paths = set(storage_map.keys()) | set(firestore_map.keys())
        
        for path in sorted(all_paths):
            row = self.results_table.rowCount()
            self.results_table.insertRow(row)
            
            filename = os.path.basename(path)
            
            in_storage = path in storage_map
            in_firestore = path in firestore_map
            
            if in_storage and in_firestore:
                status = "✅ OK"
                status_color = QColor("#15803D")
            elif in_storage and not in_firestore:
                status = "⚠️ Huérfano"
                status_color = QColor("#DC2626")
            else:
                status = "❌ Faltante"
                status_color = QColor("#EA580C")
            
            # Nombre
            item_name = QTableWidgetItem(filename)
            self.results_table.setItem(row, 0, item_name)
            
            # Factura
            invoice_num = "N/A"
            if in_firestore:
                invoice_num = firestore_map[path].get('invoice_number', 'N/A')
            item_invoice = QTableWidgetItem(invoice_num)
            self.results_table.setItem(row, 1, item_invoice)
            
            # Estado
            item_status = QTableWidgetItem(status)
            item_status.setForeground(status_color)
            font = item_status.font()
            font.setBold(True)
            item_status.setFont(font)
            self.results_table.setItem(row, 2, item_status)
            
            # Tamaño
            size_str = "N/A"
            if in_storage:
                size_kb = storage_map[path]['size'] / 1024
                size_str = f"{size_kb:.1f} KB"
            item_size = QTableWidgetItem(size_str)
            item_size.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.results_table.setItem(row, 3, item_size)
            
            # Fecha
            date_str = "N/A"
            if in_storage:
                updated = storage_map[path]['updated']
                if updated:
                    date_str = updated.strftime("%Y-%m-%d %H:%M")
            item_date = QTableWidgetItem(date_str)
            self.results_table.setItem(row, 4, item_date)
            
            # Ruta
            item_path = QTableWidgetItem(path)
            item_path.setForeground(QColor("#64748B"))
            self.results_table.setItem(row, 5, item_path)
    
    # ========================================
    # UTILIDADES
    # ========================================
    
    def _log(self, message: str):
        """Agrega mensaje al log."""
        self.log_text.append(message)
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
    
    def _clear_results(self):
        """Limpia los resultados."""
        self.results_table.setRowCount(0)
        self.log_text.clear()
        
        for stat in [self.stat_storage, self.stat_firestore, self.stat_match, 
                     self.stat_orphan, self.stat_missing]:
            stat["value"].setText("0")
    
    def _export_excel(self):
        """Exporta resultados a Excel."""
        if self.results_table.rowCount() == 0:
            QMessageBox.warning(self, "Sin Datos", "No hay resultados para exportar.")
            return
        
        fname, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar Reporte",
            f"Auditoria_Adjuntos_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not fname:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Auditoría Adjuntos"
            
            headers = ["Archivo", "Factura", "Estado", "Tamaño", "Última Modificación", "Ruta"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            
            for row in range(self.results_table.rowCount()):
                for col in range(self.results_table.columnCount()):
                    item = self.results_table.item(row, col)
                    if item:
                        ws.cell(row + 2, col + 1, item.text())
            
            wb.save(fname)
            
            QMessageBox.information(self, "Éxito", f"Reporte exportado a:\n{fname}")
            
        except ImportError:
            QMessageBox.warning(self, "Error", "Instala openpyxl:\n\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error exportando:\n{e}")

    def _investigate_orphans(self):
        """Investiga facturas huérfanas buscando en toda la base de datos."""
        if not self.attachments_storage or not self.db:
            QMessageBox.warning(self, "Sin Datos", "Primero escanea adjuntos.")
            return
        
        # Filtrar solo huérfanos
        firestore_paths = set(att['storage_path'] for att in self.attachments_firestore)
        orphans = [att for att in self.attachments_storage if att['name'] not in firestore_paths]
        
        if not orphans:
            QMessageBox.information(self, "Sin Huérfanos", "No hay adjuntos huérfanos.")
            return
        
        self._log("")
        self._log("="*80)
        self._log(f"🔎 INVESTIGANDO {len(orphans)} ADJUNTOS HUÉRFANOS")
        self._log("="*80)
        
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(len(orphans))
        self.progress_bar.setValue(0)
        
        QApplication.processEvents()
        
        found = []
        not_found = []
        
        for idx, orphan in enumerate(orphans, 1):
            self.progress_bar.setValue(idx)
            QApplication.processEvents()
            
            filename = os.path.basename(orphan['name'])
            self._log(f"\n🔍 {idx}/{len(orphans)}: {filename}")
            
            # Extraer número de factura del nombre del archivo
            # Formato: E310001906913_101008172.jpeg
            invoice_number = filename.split('_')[0] if '_' in filename else filename.split('.')[0]
            
            self._log(f"   Buscando factura: {invoice_number}")
            
            try:
                # Buscar en TODAS las facturas (sin filtro de empresa)
                invoices_ref = self.db.collection('invoices')
                query = invoices_ref.where('invoice_number', '==', invoice_number).limit(1)
                docs = list(query.stream())
                
                if docs:
                    doc = docs[0]
                    data = doc.to_dict()
                    
                    company_id = data.get('company_id', 'N/A')
                    invoice_date = data.get('invoice_date', 'N/A')
                    storage_path = data.get('attachment_storage_path') or data.get('storage_path', 'N/A')
                    
                    self._log(f"   ✅ FACTURA ENCONTRADA:")
                    self._log(f"      • ID: {doc.id}")
                    self._log(f"      • Empresa: {company_id}")
                    self._log(f"      • Fecha: {invoice_date}")
                    self._log(f"      • Storage Path: {storage_path}")
                    self._log(f"      • Ruta Real: {orphan['name']}")
                    
                    if storage_path != orphan['name']:
                        self._log(f"      ⚠️ DESAJUSTE: Las rutas NO coinciden")
                        
                        found.append({
                            'filename': filename,
                            'invoice_id': doc.id,
                            'invoice_number': invoice_number,
                            'company_id': company_id,
                            'storage_path_db': storage_path,
                            'storage_path_real': orphan['name'],
                            'mismatch': True
                        })
                    else:
                        found.append({
                            'filename': filename,
                            'invoice_id': doc.id,
                            'invoice_number': invoice_number,
                            'company_id': company_id,
                            'mismatch': False
                        })
                else:
                    self._log(f"   ❌ No se encontró factura con número: {invoice_number}")
                    
                    # Intentar búsqueda parcial
                    partial_results = []
                    try:
                        # Buscar facturas que contengan parte del número
                        all_invoices = invoices_ref.stream()
                        for inv_doc in all_invoices:
                            inv_data = inv_doc.to_dict()
                            inv_num = inv_data.get('invoice_number', '')
                            if invoice_number in inv_num or inv_num in invoice_number:
                                partial_results.append({
                                    'id': inv_doc.id,
                                    'number': inv_num,
                                    'company': inv_data.get('company_id', 'N/A')
                                })
                                if len(partial_results) >= 5:
                                    break
                        
                        if partial_results:
                            self._log(f"   🔍 Facturas similares encontradas:")
                            for pr in partial_results:
                                self._log(f"      • {pr['number']} (ID: {pr['id']}, Empresa: {pr['company']})")
                    except:
                        pass
                    
                    not_found.append({
                        'filename': filename,
                        'invoice_number': invoice_number,
                        'storage_path': orphan['name']
                    })
            
            except Exception as e:
                self._log(f"   ❌ Error: {e}")
        
        # Resumen
        self._log("")
        self._log("="*80)
        self._log("📊 RESUMEN DE INVESTIGACIÓN")
        self._log("="*80)
        self._log(f"   Total huérfanos investigados: {len(orphans)}")
        self._log(f"   ✅ Facturas encontradas: {len(found)}")
        self._log(f"   ❌ Facturas NO encontradas: {len(not_found)}")
        
        if found:
            mismatches = sum(1 for f in found if f.get('mismatch'))
            self._log(f"   ⚠️ Con desajuste de ruta: {mismatches}")
        
        self.progress_bar.setVisible(False)
        
        # Ofrecer corrección
        if found and any(f.get('mismatch') for f in found):
            reply = QMessageBox.question(
                self,
                "Corregir Rutas",
                f"Se encontraron {sum(1 for f in found if f.get('mismatch'))} facturas con rutas incorrectas.\n\n"
                f"¿Deseas corregir las rutas automáticamente en Firestore?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self._fix_mismatched_paths(found)


    def _fix_mismatched_paths(self, mismatched_invoices: list):
        """Corrige las rutas de storage_path en Firestore."""
        self._log("")
        self._log("="*80)
        self._log("🔧 CORRIGIENDO RUTAS DE ADJUNTOS")
        self._log("="*80)
        
        fixed = 0
        failed = 0
        
        for inv in mismatched_invoices:
            if not inv.get('mismatch'):
                continue
            
            try:
                invoice_id = inv['invoice_id']
                correct_path = inv['storage_path_real']
                
                self._log(f"\n🔧 Corrigiendo {inv['invoice_number']}...")
                self._log(f"   Nueva ruta: {correct_path}")
                
                # Actualizar en Firestore
                doc_ref = self.db.collection('invoices').document(invoice_id)
                doc_ref.update({
                    'attachment_storage_path': correct_path
                })
                
                fixed += 1
                self._log(f"   ✅ Corregido")
                
            except Exception as e:
                failed += 1
                self._log(f"   ❌ Error: {e}")
        
        self._log("")
        self._log(f"📊 RESULTADO:")
        self._log(f"   ✅ Corregidos: {fixed}")
        self._log(f"   ❌ Fallidos: {failed}")
        
        QMessageBox.information(
            self,
            "Corrección Completa",
            f"Rutas corregidas: {fixed}\n"
            f"Fallidos: {failed}\n\n"
            f"Vuelve a escanear para verificar."
        )

def main():
    """Función principal."""
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 9))
    
    window = AttachmentAuditor()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()